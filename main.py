import tkinter
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font
import copy
from openpyxl.drawing.image import Image
class EasyUpper:
    def __init__(self):
        self.root = Tk()
        self.init()
        self.file_panel = Label(self.root, height=1, width=10)
    def init(self):
        self.root.geometry("250x300")
        self.root.geometry("+700+400")
        self.root.title('Combiner V2.0')
        self.root.resizable(False, False)
        self.root.configure(bg='gray')
        # self.enabled = print(1)
    def run(self):
        self.draw_menu()
        self.draw_widgets()
        self.root.mainloop()
    def draw_menu(self):
        menu_bar = Menu(self.root)
        file_menu = Menu(menu_bar, tearoff=0)
        open_sub_menu = Menu(file_menu, tearoff=0)
        open_sub_menu.add_command(label='File', command=self.open_excel_file1)
        open_sub_menu.add_command(label="Database", command=self.open_excel_database)
        file_menu.add_cascade(label="Open", menu=open_sub_menu)
        file_menu.add_command(label="Format DB", command=self.format_db)
        # file_menu.add_command(label="Languages", command=self.landguages)
        menu_bar.add_cascade(label="File", menu=file_menu)
        self.root.configure(menu=menu_bar)
    def draw_widgets(self):
        self.root.center_button  = tkinter.Button(self.root, text="Center", command=self.center , fg='Black', bg='Orange',  height= 4, width=10)
        self.root.center_button.place(x=15, y=70)
        self.root.font_button = tkinter.Button(self.root, text="Font_size", command=self.font_size, fg='Black', bg='Orange', height=2, width=10)
        self.root.font_button.place(x=135, y=195)
        # self.root.format_button = tkinter.Button(self.root, text="Format database", command=self.center, fg='Black', bg='Orange', height=4, width=10)
        # self.root.format.place(x=15, y=70)
        self.root.start_button = tkinter.Button(self.root, text="Start!\n UPPER", command=self.load_UP , fg='Black', bg='Orange', height= 2, width=10)
        self.root.start_button.place(x=135, y=98)

        self.root.to_data_button = tkinter.Button(self.root, text=" Transfer \n Values", command=self.autocomplete , fg='Black', bg='Orange', height= 4, width=10)
        self.root.to_data_button.place(x=15, y=165)

        self.root.text = tkinter.Text(self.root, height=1 , width=11, font=('Arrial', 10))
        self.root.text.place(x=133, y=75)

        self.root.label_col = tkinter.Label(self.root, text="Select Column\n    (Number): ", font=('Arrial', 10), fg=("Orange"), bg='Gray')
        self.root.label_col.place(x=127, y=155)

        self.root.label_col = tkinter.Label(self.root, text="        File : ", font=('Arrial', 10), fg=("Orange"), bg='Gray')
        self.root.label_col.place(x=7, y=8)

        self.root.label_col = tkinter.Label(self.root, text="Database : ", font=('Arrial', 10), fg=("Orange"), bg='Gray')
        self.root.label_col.place(x=5, y=30)
    def open_excel_file1(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*xlsx")])
        self.file_panel = Label(self.root, text=file_path[25:50:])
        if  self.file_panel == file_path:
            self.workbook.close
        self.file_panel.file = file_path
        self.file_panel.place(x=75, y= 10)
        self.file_panel = self.file_panel
        self.workbook = openpyxl.load_workbook(self.file_panel.file)
        try:
          s1 = self.workbook.active
        except PermissionError:
            messagebox.showinfo('error', 'File open in another programm')
        merged_cells_copy = copy.copy(s1.merged_cells.ranges)
        for merged_cell_range in merged_cells_copy:
            s1.unmerge_cells(str(merged_cell_range))
        self.workbook.save(self.file_panel.file)
        for image in s1._images:
            s1._images.remove(image)
            self.workbook.save(self.file_panel.file)
    def open_excel_database(self):
        file_path_data = filedialog.askopenfilename(filetypes=[("Excel files", "*xlsx")])
        self.file_panel_data = Label(self.root, text=file_path_data[25:50:])
        if  self.file_panel_data == file_path_data:
            self.workbook_data.close
        self.file_panel_data.file = file_path_data
        self.file_panel_data.place(x=75, y= 32)
        self.file_panel_data = self.file_panel_data
        self.workbook_data = openpyxl.load_workbook(self.file_panel_data.file)
        self.sheet_data = self.workbook_data.active

        s2 = self.workbook_data.active
        merged_cells_copy = copy.copy(s2.merged_cells.ranges)
        for merged_cell_range in merged_cells_copy:
            s2.unmerge_cells(str(merged_cell_range))
        try:
            self.workbook_data.save(self.file_panel_data.file)
        except PermissionError:
            messagebox.showinfo('error', 'File open in another programm')
        for image in s2._images:
            s2._images.remove(image)
            self.workbook_data.save(self.file_panel_data.file)
    def load_UP(self):
        from openpyxl.utils import cell
        try:
            self.my_number = int(self.root.text.get("1.0", "end-1c"))
        except ValueError:
            messagebox.showinfo('Error', 'Select Column')
        else:
            pass
        try:
            self.col_letter = cell.get_column_letter(self.my_number)
        except AttributeError:
            messagebox.showinfo('Error', 'Select Column')  #
        if self.file_panel:
            try:
                self.sheet = self.workbook.active
            except AttributeError:
                messagebox.showinfo('Error', '     Open file!\n  File -> Open -> File')
            column_data = []
            for cell in self.sheet[self.col_letter]:
                if cell.value == None:
                    column_data.append('')
                else:
                    column_data.append(cell.value)
        new_list = []
        for item in column_data:
            if item.strip() == None:
                new_list.append('')
            else:
                new_list.append(item.upper())

        for index, item in enumerate(new_list):
                cell = self.sheet.cell(row= index +1, column=self.my_number)
                cell.value = item
        try:
            self.workbook.save(self.file_panel.file)
        except PermissionError:
            messagebox.showinfo('Error', 'File open in another programm')
        else:
            self.workbook.close()
            messagebox.showinfo('Complete', 'Saved!')
    def center(self):
        try:
            self.sheet = self.workbook.active
        except AttributeError:
            messagebox.showinfo('Error', '     Open file!\n  File -> Open -> File')

        cell_range = self.sheet['A1':'L200']
        from openpyxl.styles import Alignment
        for row in cell_range:
            for cell in row:
                align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.alignment = align
        try:
            self.workbook.save(self.file_panel.file)
        except PermissionError:
            messagebox.showinfo('Error', 'File open in another programm')
        else:
            self.workbook.close()
            messagebox.showinfo('Complete', 'Saved!')
    def format_db(self):
        try:
            s2 = self.workbook_data.active
        except AttributeError:
            messagebox.showinfo('Error', '     Add Database file!\n  File -> Open -> Database')  #

        plates_list = []
        plates_strip = []

        for row in s2.iter_rows():
            if row is not None:
                plates = row[3].value
                plates_list.append(plates)

        for i, plate in enumerate(plates_list):
            if plate is not None:
                stripped_plate = plate.rstrip().replace(' ', '')
                plates_strip.append(stripped_plate)
            else:
                plates_strip.append('')

        for row, stripped_plate in zip(s2.iter_rows(), plates_strip):

            row_index = row[8].row
            cell2 = s2.cell(row=row_index, column=4)
            cell2.value = stripped_plate

        image = Image('LG_Title.png')
        try:
            s2.add_image(image, 'B1')
        except ValueError:
            pass

        try:
            self.workbook_data.save(self.file_panel_data.file)

            self.workbook_data.close()
        except PermissionError:
            messagebox.showinfo('Error', 'File open in another programm')
        else:

            messagebox.showinfo('Complete', 'Saved!')  #
    def autocomplete(self):
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        from openpyxl.utils import column_index_from_string
        from PIL import ImageChops

        try:
            s1 = self.workbook.active
        except AttributeError:
            messagebox.showinfo('Error', '     Open file!\n  File -> Open -> File')
        try:
            s2 = self.workbook_data.active
        except AttributeError:
            messagebox.showinfo('Error', '     Add Database file!\n  File -> Open -> Database')  #

        plate_list = []
        plate_strip = []
        names_list = []
        places_list = []
        control_list = []
        card_list = []

        card_db = []
        correct_card = []
        plate_db = []
        plate_replace = []
        names_db = []
        places_db = []
        control_db = []
        for row in s1.iter_rows():

            if row is not None:
                plate = row[6].value
                names = row[11].value
                places = row[8].value
                control = row[10].value
                card = row[3].value
                card_list.append(card)
                control_list.append(control)
                plate_list.append(plate)
                names_list.append(names)
                places_list.append(places)

        for row in s2.iter_rows():
            if row is not None:
                plates_data = row[3].value
                names_data = row[8].value
                places_data = row[5].value
                control_data = row[7].value
                card_data = row[0].value
                card_db.append(card_data)
                control_db.append(control_data)
                places_db.append(places_data)
                plate_db.append(plates_data)
                names_db.append(names_data)

        for i in card_db:
            if i == 'DEPARTMENT LEADER':
                correct_card.append('TOP MANAGEMENT')
            elif i == 'DIRECTOR':
                correct_card.append('TOP MANAGEMENT')
            else:
                correct_card.append(i)

        skip_spaces = True
        for i in plate_list:
            if i is not None:
                if skip_spaces:
                    plate_strip.append(i)
                    skip_spaces = False
                else:
                    plate_strip.append(i.replace(' ', ''))

        for i in plate_db:
            if i is not None:
                if skip_spaces:
                    plate_replace.append(i)
                    skip_spaces = False
                else:
                    plate_replace.append(i.replace(' ', ''))

        for row1 in s1.iter_rows():
            plate1 = row1[6].value

            if plate1 in plate_db:
                index = plate_db.index(plate1)
                value = names_db[index]
                row_index = row1[8].row
                cell = s1.cell(row=row_index, column=12)
                if cell.value is None:
                    cell.value = value

                control_value = control_db[index]
                row_index = row1[8].row
                cell2 = s1.cell(row=row_index, column=11)
                if cell2.value is None:
                    cell2.value = control_value

                card_value = correct_card[index]
                row_index = row1[8].row
                cell3 = s1.cell(row=row_index, column=4)
                if cell3.value is None:
                    cell3.value = card_value

                value_places = places_db[index]
                row_index = row1[8].row
                cell1 = s1.cell(row=row_index, column=9)
                if cell1.value is None:
                    try:
                        cell1.value = value_places
                    except AttributeError:
                        messagebox.showinfo('Error', 'File have a merged cells')

        for row in s1.iter_rows(min_row=8, max_row=100):
            row_index = row[8].row
            cell = s1.cell(row=row_index , column=6)
            if cell.value is None:
                cell.value = 'LG'
            cell1 = s1.cell(row=row_index, column=5)
            if cell1.value is None:
                cell1.value = 'OSOBOWY'

        try:
            image = Image('LG_Title.png')
            s1.add_image(image, 'B1')
        except FileNotFoundError:
            pass

        try:
            self.workbook.save(self.file_panel.file)

        except PermissionError:
            messagebox.showinfo('Error', 'File open in another programm')
        else:
            messagebox.showinfo('Complete', 'Saved!') #
        self.workbook.close()
    def landguages(self):
        current_language = 'ru'
        translations = {
            'ru': {
                'file_text': 'Файл',
                'db_text': 'База данных',
                'open_text': 'Открыть',
                'help_text': 'Помощь',
                'languages_text': 'Язык',
                'transfer_button_text': 'Перенос \n значений',
                'select_text': 'Выберите столбец \n     (номер)',
                'center_text': 'Центрировать \n  все ячейки',
                'start_text': '   Запуск \n смена регистра',
                'path_text': 'Путь к файлу: ',
                'path_db_text': 'Путь к \n Базе данных',
                'error_perm': 'Ошибка доступа. Файл открыт в другой программе!',
                'value_err': 'Ошибка значения'
            },
            'en': {
                'file_text': 'File',
                'db_text': 'Database',
                'open_text': 'Open',
                'help_text': 'Help',
                'languages_text': 'Languages',
                'transfer_button_text': 'Transfer \n values',
                'select_text': 'Select column \n     (number)',
                'center_text': 'Center \n all cells',
                'start_text': '   Start \n change case',
                'path_text': 'File path: ',
                'path_db_text': 'Database \n path',
                'error_perm': 'Access error. File is open in another program!',
                'value_err': 'Value error'
            }
        }
        # self.draw_widgets(translations)
        self.draw_menu(translations)
        self.draw_menu(current_language)
    def font_size(self):
        try:
            ws = self.workbook.active
        except AttributeError:
            messagebox.showinfo('Error', '     Open file!\n  File -> Open -> File')
        font = Font(size = 11)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
        try:
            self.workbook.save(self.file_panel.file)
        except PermissionError:
            messagebox.showinfo('Error', 'File open in another programm')
        else:
            self.workbook.close()
            messagebox.showinfo('Complete', 'Saved!')
if __name__ =="__main__":
    EasyUpper().run()

#skflsdh;fsldhgsdGhsdklgs
#залупка




# реализовать перевод каким нибудь способом! хотя бы ru/pl в идеале ru/pl/en
# провести дебаг сессию на стабильность(), вычистить код
# подготовка к релизу и упаковка в exe файл(бета версия работает стабильно!)!
# подготовка к работе с сырым рапортом
# ограничить область сканирования до 500 строк на все функции!!!



